from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb =Workbook()
dest_filename = 'empty_book.xlsx'

ws = wb.active
# 设定当前工作表标题
ws.title = 'range names'
# 创建40行
for row in range(1,40):
    # 在每一行，将0-599写入列
    ws.append(range(600))
# 创建一张新表
ws2 = wb.create_sheet(title='Pi')
# F5单元格中写入3.14
ws2["F5"] = 3.14

ws3 = wb.create_sheet(title='Data')
for row in range(10, 20):
    for col in range(27,54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
wb.save(filename=dest_filename)
