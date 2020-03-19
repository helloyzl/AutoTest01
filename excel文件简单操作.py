from openpyxl import Workbook
import os, datetime

wb = Workbook()
ws =wb.active
# ws['A1'] = 42
ws.append([1,2,3])

# ws['A2'] = datetime.datetime.now()
# 打印当前路径
print(os.getcwd())
# 将文件保存在当前路径
wb.save('sample.xlsx')