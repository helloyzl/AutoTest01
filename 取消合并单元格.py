from openpyxl import Workbook
from openpyxl.utils import get_column_interval

wb = Workbook()
ws = wb.active
ws.title = '测试表1'
# 第1-9行
k = 0
for i in range(1,10):
    # k = 1
    # 20-29列
    for j in range(1,10):  # 先循环最里层的
        _ = ws.cell(row=i,column=j,value=k)
        k += 1
# 合并A2：D2的单元格
# ws.merge_cells('A2:D2')
# print(ws['B2'].value)
# 取消合并A2：D2的单元格
# ws.unmerge_cells("A2:D2")

wb.save(filename='test.xlsx')