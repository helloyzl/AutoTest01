from openpyxl import load_workbook


def do_xlsx():
    wb = load_workbook(filename=r'C:\Users\Administrator\Desktop\商品导入模板.xlsx')
    # 打开指定sheet表
    ws = wb['商品下载模板']
    # 打印“商品下载模板表的A1的值”
    print(ws['A1'].value)
    # 循环9次
    # for x in range(9):
    #     k = get_ran_good_name()
    k = get_ran_good_name()
    print(k)
    print(type(k))

    for i in range(2,11):
        # k = get_ran_good_name()  # 商品名  只会打印 测试商品1
        # 写多少列
        for j in range(1, 2):
            _ = ws.cell(row=i, column=j, value="{0}".format(k[i]))

    # 写项目列
    for i2 in range(2,11):
        # 写多少列
        for j2 in range(2, 3):
            _ = ws.cell(row=i2, column=j2, value=4)

    # 写品牌名称
    for i3 in range(2,11):
        # 写多少列
        for j3 in range(3, 4):
            _ = ws.cell(row=i3, column=j3, value='测试品牌')


    # 写基本单位名称
    for i4 in range(2,11):
        # 写多少列
        for j4 in range(4, 5):
            _ = ws.cell(row=i4, column=j4, value='包')


    # 写商品保质期数量
    for i5 in range(2,11):
        # 写多少列
        for j5 in range(6, 7):
            _ = ws.cell(row=i5, column=j5, value=1)  # 保质期都是1


    # 写商品保质期单位，都设定为月
    for i6 in range(2,11):
        # 写多少列
        for j6 in range(7, 8):
            _ = ws.cell(row=i6, column=j6, value='月')  # 保质期单位都是月


    # 写商品进项税率，都设定为0
    for i7 in range(2,11):
        # 写多少列
        for j7 in range(8, 9):
            _ = ws.cell(row=i7, column=j7, value=0)  # 商品进项税率0

    # 写商品销项税率，都设定为0
    for i8 in range(2,11):
        # 写多少列
        for j8 in range(9, 10):
            _ = ws.cell(row=i8, column=j8, value=0)  # 商品销项税率0


    # 保存才会执行的结果
    wb.save(filename='test2.xlsx')


def get_ran_good_name():
    name_lst = []
    for x in range(1,10):
        print('测试商品'+ str(x))
        name_lst.append('测试商品'+ str(x))
    return name_lst


if __name__ == '__main__':
    do_xlsx()
    # x = get_ran_good_name()